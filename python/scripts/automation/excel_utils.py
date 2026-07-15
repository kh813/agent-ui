"""Excel ファイル操作ユーティリティ

ファイル操作・データ読み書き・集計など、Excel 自動化に必要な
共通操作をまとめたモジュール。chrome_utils.py と組み合わせて使うことを想定している。

【関数一覧】
  ファイル : open_workbook, new_workbook, open_or_create, save_workbook
  シート   : get_sheet, get_or_create_sheet, list_sheets
  読み取り : read_all, read_column, find_row, get_last_row
  書き込み : write_cell, append_row, append_rows, update_row
  集計     : sum_column, count_column, filter_rows, aggregate
  変換     : from_records, to_records, from_csv

使用例（Web取得データをExcelに追記）:
    from chrome_utils import open_url, get_structured_list
    from excel_utils import open_or_create, get_or_create_sheet, append_rows, save_workbook

    records = get_structured_list(page, "tr.data-row", {
        "date": "td.date", "value": "td.value"
    })
    wb = open_or_create("~/Downloads/data.xlsx")
    sheet = get_or_create_sheet(wb, "Sheet1")
    append_rows(sheet, records)
    save_workbook(wb, "~/Downloads/data.xlsx")
"""

from __future__ import annotations
import csv
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def _resolve(path: str | Path) -> Path:
    return Path(path).expanduser().resolve()


def _col_index(sheet: Worksheet, col: int | str) -> int:
    """列番号（1始まり）または列名（ヘッダー文字列）を列インデックスに変換する。"""
    if isinstance(col, int):
        return col
    for cell in sheet[1]:
        if cell.value == col:
            return cell.column
    raise ValueError(f"列が見つかりません: {col!r}")


def _header_map(sheet: Worksheet) -> dict[str, int]:
    """1行目をヘッダーとして {列名: 列インデックス} を返す。"""
    return {
        cell.value: cell.column
        for cell in sheet[1]
        if cell.value is not None
    }


# ------------------------------------------------------------------
# ファイル操作
# ------------------------------------------------------------------

def open_workbook(path: str | Path) -> Workbook:
    """既存の Excel ファイルを開く。"""
    return load_workbook(_resolve(path))


def new_workbook() -> Workbook:
    """新規ワークブックを作成する。デフォルトシート名は "Sheet1"。"""
    wb = Workbook()
    wb.active.title = "Sheet1"
    return wb


def open_or_create(path: str | Path) -> Workbook:
    """ファイルが存在すれば開き、なければ新規作成する。"""
    p = _resolve(path)
    return load_workbook(p) if p.exists() else new_workbook()


def save_workbook(wb: Workbook, path: str | Path) -> Path:
    """ワークブックを指定パスに保存する。保存先パスを返す。"""
    p = _resolve(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    wb.save(p)
    print(f"保存しました: {p}")
    return p


# ------------------------------------------------------------------
# シート操作
# ------------------------------------------------------------------

def get_sheet(wb: Workbook, name: str) -> Worksheet:
    """シート名でシートを取得する。存在しなければ KeyError を送出する。"""
    if name not in wb.sheetnames:
        raise KeyError(f"シートが見つかりません: {name!r}（存在: {wb.sheetnames}）")
    return wb[name]


def get_or_create_sheet(wb: Workbook, name: str) -> Worksheet:
    """シートが存在すれば取得し、なければ作成して返す。"""
    if name in wb.sheetnames:
        return wb[name]
    return wb.create_sheet(name)


def list_sheets(wb: Workbook) -> list[str]:
    """ワークブック内のシート名一覧を返す。"""
    return wb.sheetnames


# ------------------------------------------------------------------
# データ読み取り
# ------------------------------------------------------------------

def read_all(sheet: Worksheet) -> list[dict]:
    """1行目をヘッダーとして、全行を辞書リストで返す。

    chrome_utils.get_structured_list() と同じ形式で返すため、両者を組み合わせやすい。
    """
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
    return [dict(zip(headers, row)) for row in rows[1:]]


def read_column(sheet: Worksheet, col: int | str) -> list:
    """指定列のデータを（ヘッダーを除く）リストで返す。

    col は列番号（1始まり）またはヘッダー文字列で指定する。
    """
    idx = _col_index(sheet, col)
    return [sheet.cell(row=r, column=idx).value for r in range(2, sheet.max_row + 1)]


def find_row(sheet: Worksheet, col: int | str, value: Any) -> int | None:
    """指定列の値が value に一致する最初の行番号（1始まり）を返す。見つからなければ None。

    update_row と組み合わせて「キー列で行を特定して更新する」パターンで使う。
    """
    idx = _col_index(sheet, col)
    for r in range(2, sheet.max_row + 1):
        if sheet.cell(row=r, column=idx).value == value:
            return r
    return None


def get_last_row(sheet: Worksheet) -> int:
    """データが入っている最終行番号を返す。空シートは 1（ヘッダー行）を返す。

    append_row / append_rows の追記位置を求めるために使う。
    """
    return sheet.max_row if sheet.max_row else 1


# ------------------------------------------------------------------
# データ書き込み
# ------------------------------------------------------------------

def write_cell(sheet: Worksheet, row: int, col: int | str, value: Any) -> None:
    """指定セルに値を書き込む。col は列番号またはヘッダー文字列。"""
    sheet.cell(row=row, column=_col_index(sheet, col), value=value)


def append_row(sheet: Worksheet, data: list | dict) -> int:
    """最終行の次に1行追加する。追記した行番号を返す。

    data がリストの場合はそのまま列順に書き込む。
    data が辞書の場合はヘッダー行と照合して列位置を合わせる。
    """
    if isinstance(data, dict):
        hmap = _header_map(sheet)
        if not hmap:
            # ヘッダーがなければキーを1行目に書いてから追記
            for i, key in enumerate(data.keys(), 1):
                sheet.cell(row=1, column=i, value=key)
            hmap = _header_map(sheet)
        row_num = get_last_row(sheet) + 1
        for key, val in data.items():
            if key in hmap:
                sheet.cell(row=row_num, column=hmap[key], value=val)
    else:
        row_num = get_last_row(sheet) + 1
        for i, val in enumerate(data, 1):
            sheet.cell(row=row_num, column=i, value=val)
    return row_num


def append_rows(sheet: Worksheet, records: list[dict | list]) -> None:
    """chrome_utils の取得結果をそのまま一括追記する。

    records が辞書リストの場合、ヘッダーが未設定なら1行目に自動生成する。
    """
    if not records:
        return
    if isinstance(records[0], dict):
        hmap = _header_map(sheet)
        if not hmap:
            keys = list(records[0].keys())
            for i, key in enumerate(keys, 1):
                sheet.cell(row=1, column=i, value=key)
            hmap = {key: i + 1 for i, key in enumerate(keys)}
        start = get_last_row(sheet) + 1
        for r_offset, record in enumerate(records):
            for key, val in record.items():
                if key in hmap:
                    sheet.cell(row=start + r_offset, column=hmap[key], value=val)
    else:
        for row in records:
            append_row(sheet, row)


def update_row(sheet: Worksheet, key_col: int | str,
               key_value: Any, updates: dict) -> bool:
    """key_col の値が key_value に一致する行を updates の内容で更新する。

    更新できた場合は True、行が見つからなければ False を返す。
    """
    row_num = find_row(sheet, key_col, key_value)
    if row_num is None:
        return False
    hmap = _header_map(sheet)
    for key, val in updates.items():
        if key in hmap:
            sheet.cell(row=row_num, column=hmap[key], value=val)
    return True


# ------------------------------------------------------------------
# 集計
# ------------------------------------------------------------------

def sum_column(sheet: Worksheet, col: int | str) -> float:
    """指定列の数値合計を返す。非数値セルは無視する。"""
    values = read_column(sheet, col)
    return sum(v for v in values if isinstance(v, (int, float)))


def count_column(sheet: Worksheet, col: int | str) -> int:
    """指定列の非空白セル数を返す。"""
    return sum(1 for v in read_column(sheet, col) if v is not None and v != "")


def filter_rows(sheet: Worksheet, col: int | str, value: Any) -> list[dict]:
    """指定列の値が value に一致する行を辞書リストで返す。"""
    idx = _col_index(sheet, col)
    headers = [cell.value for cell in sheet[1]]
    result = []
    for r in range(2, sheet.max_row + 1):
        row_vals = [sheet.cell(row=r, column=c).value for c in range(1, len(headers) + 1)]
        if row_vals[idx - 1] == value:
            result.append(dict(zip(headers, row_vals)))
    return result


def aggregate(sheet: Worksheet, group_by: str,
              operations: dict[str, str]) -> list[dict]:
    """group_by 列でグループ化し、各列に指定した集計を適用して辞書リストで返す。

    operations = {"売上": "sum", "件数": "count", "単価": "mean"}
    対応する集計: "sum" / "count" / "mean" / "max" / "min"

    使用例:
        result = aggregate(sheet, "部署", {"売上": "sum", "案件数": "count"})
        # → [{"部署": "営業部", "売上": 1500000, "案件数": 12}, ...]
    """
    import pandas as pd

    records = read_all(sheet)
    if not records:
        return []

    df = pd.DataFrame(records)
    if group_by not in df.columns:
        raise ValueError(f"列が見つかりません: {group_by!r}")

    agg_map = {col: op for col, op in operations.items() if col in df.columns}
    grouped = df.groupby(group_by, as_index=False).agg(agg_map)
    return grouped.to_dict(orient="records")


# ------------------------------------------------------------------
# 変換
# ------------------------------------------------------------------

def from_records(sheet: Worksheet, records: list[dict],
                 start_row: int = 1) -> None:
    """辞書リストをシートの start_row 行目から書き込む（既存データは上書き）。

    append_rows との違い: 先頭から書き直す場合に使う。
    ヘッダーは records[0] のキーから自動生成する。
    """
    if not records:
        return
    keys = list(records[0].keys())
    for i, key in enumerate(keys, 1):
        sheet.cell(row=start_row, column=i, value=key)
    for r_offset, record in enumerate(records, 1):
        for i, key in enumerate(keys, 1):
            sheet.cell(row=start_row + r_offset, column=i, value=record.get(key))


def to_records(sheet: Worksheet) -> list[dict]:
    """シート全体を辞書リストとして返す。read_all の別名。"""
    return read_all(sheet)


def from_csv(sheet: Worksheet, csv_path: str | Path) -> None:
    """CSV ファイルの内容をシートに読み込む。既存データはすべて消去する。

    ヘッダー行を含む CSV を想定している。
    """
    p = _resolve(csv_path)
    for row in sheet.iter_rows():
        for cell in row:
            cell.value = None

    with open(p, newline="", encoding="utf-8-sig") as f:
        for r_offset, row in enumerate(csv.reader(f), 1):
            for c_offset, val in enumerate(row, 1):
                sheet.cell(row=r_offset, column=c_offset, value=val)
